"""importing data from a excel file"""

import openpyxl
import sys 
import argparse

class Import:
    """Import data from excel file
    
    Attributes:
        path(string): path of the excel fiel
        file(object)" : workbook active sheet object
    """

    def __init__(self, path):
        """Import data from excel file
    
        Attributes:
        path(string): path of the excel file
        file(object)" : workbook active sheet object

        Args:
            path(string): path of the excel file
        """

        self.path = path
        work_book = openpyxl.load_workbook(self.path)
        self.file = work_book.active
        
    
    def examine_data(self):
        """Save all the data to a dictionary
        
        Returns:
            dic(dictionary): the dictionary contains all the data
        """
        dic = {}
        max_row = self.file.max_row
        
        #Company name
        for i in range(3,max_row+1):
            if self.file.cell(row = 2, column = 1).value in dic.keys():
                dic[self.file.cell(row = 2, column = 1).value] \
                    .append(self.file.cell(row = i, column = 1).value)
            else:
                dic[self.file.cell(row = 2, column = 1).value] = []
                dic[self.file.cell(row = 2, column = 1).value] \
                    .append(self.file.cell(row = i, column = 1).value)
        
        #Account
        for i in range(3,max_row+1):
            if self.file.cell(row = 2, column = 2).value in dic.keys():
                dic[self.file.cell(row = 2, column = 2).value] \
                    .append(self.file.cell(row = i, column = 2).value)
            else:
                dic[self.file.cell(row = 2, column = 2).value] = []
                dic[self.file.cell(row = 2, column = 2).value] \
                    .append(self.file.cell(row = i, column = 2).value)

        #Due date
        for i in range(3,max_row+1):
            if self.file.cell(row = 2, column = 3).value in dic.keys():
                dic[self.file.cell(row = 2, column = 3).value] \
                    .append(self.file.cell(row = i, column = 3).value)
            else:
                dic[self.file.cell(row = 2, column = 3).value] = []
                dic[self.file.cell(row = 2, column = 3).value] \
                    .append(self.file.cell(row = i, column = 3).value)

        #Total amount due
        for i in range(3,max_row+1):
            if self.file.cell(row = 2, column = 4).value in dic.keys():
                dic[self.file.cell(row = 2, column = 4).value] \
                    .append(self.file.cell(row = i, column = 4).value)
            else:
                dic[self.file.cell(row = 2, column = 4).value] = []
                dic[self.file.cell(row = 2, column = 4).value] \
                    .append(self.file.cell(row = i, column = 4).value)

        #Monthly Budget 
        for i in range(3,max_row+1):
            if self.file.cell(row = 2, column = 5).value in dic.keys():
                dic[self.file.cell(row = 2, column = 5).value] \
                    .append(self.file.cell(row = i, column = 5).value)
            else:
                dic[self.file.cell(row = 2, column = 5).value] = []
                dic[self.file.cell(row = 2, column = 5).value] \
                    .append(self.file.cell(row = i, column = 5).value)
        
        return dic 


def parse_args(my_args_list):
    """Create a parse_args condition
    
    Args:
        my_args_list(list): sys input list
    
    Returns:
        args: the parser object createed
    """

    parser = argparse.ArgumentParser(description='Process the path')
    parser.add_argument('file_name',type = str, help='File name')
    args = parser.parse_args(my_args_list)
    return args
    

if __name__ == '__main__':
    """Import the file and print the data""" 
    
    #new_args = parse_args(sys.argv[1:])
    #my_import = Import(sys.argv[1])
    #my_import.examine_data()
    my_import = Import("expensetracker.xlsx")
    print(my_import.examine_data())
